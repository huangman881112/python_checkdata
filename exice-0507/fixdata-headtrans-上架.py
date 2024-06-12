from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    # print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

head_inbound_exampl = """
{
	"billNo": "PAP20240415084",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-10T07:01:26.199",
	"billTypeEnum": "SELF_BUILD_INBOUND",
	"operationTypeEnum": "PUTAWAY",
	"originBillNo": "2403H-S191-CA03",
	"skuList": [
		{
			"badHoldQty": 0,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": -533,
			"inTransitQty": 0,
			"oldSkuCode": "FTPLDB-0005",
			"platform": "10001",
			"skuCode": "FTPLDB-0005",
			"store": "10415",
			"totalQty": 0,
			"useQty": 533
		}
	],
	"warehouseCode": "112"
}

"""


def init_data(rma_exampl):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        result[data_pro] = rma_exampl[data_pro]
    # print(result)
    return result


filename = "d:\\tmp\\2404H-S33-CA03.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)

# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    biz_type = row[1]
    if biz_type != 'GOOD_PUTAWAY_HOLD':
        continue
    skuCode = row[9]
    oldSkuCode = row[10]
    platform = row[7]
    store = row[8]
    billNo = row[19]
    origin_bill_no = row[18]
    bill_time = row[15]
    warehouse_code = row[6]
    qty = int(row[11])
    update_time = row[15]
    req_exam = init_data(head_inbound_exampl)
    req_exam["billNo"] = billNo
    req_exam["originBillNo"] = origin_bill_no
    req_exam["remark"] = origin_bill_no
    # print(type(update_time))
    print(update_time)
    updateTime = datetime.fromisoformat(update_time)
    # update_time = datetime.strftime(update_time,"%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    req_exam["warehouseCode"] = warehouse_code
    skuList = req_exam["skuList"]
    for sku in skuList:
        sku["platform"] = platform
        sku["store"] = store
        sku["skuCode"] = skuCode
        sku["oldSkuCode"] = oldSkuCode
        sku["holdQty"] = qty* -1
        sku["useQty"] = qty


    # req_exam = json.dumps(req_exam)
    print(req_exam)
    # 发送POST请求
    response = requests.post(url, json=req_exam, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i>=1:
    #     break
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
# workbook.save(f_tm + 'example.xlsx')
""
