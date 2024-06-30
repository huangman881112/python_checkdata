from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import commonUtil

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

billlist = """
GOC64736899
"""


req_param = """
{
	"billNo": "GOC64809697",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-23T23:40:03.043",
	"billTypeEnum": "SALE_OUTBOUND_ORDERS",
	"operationTypeEnum": "HOLD",
	"originBillNo": "GOC64809697",
	"skuList": [
		{
			"badHoldQty": 0,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": 1,
			"inTransitQty": 0,
			"oldSkuCode": "FTHKAM-0004-DS",
			"platform": "10002",
			"skuCode": "FTHKAM-0004-DS",
			"store": "11366",
			"totalQty": 0,
			"useQty": -1
		}
	],
	"warehouseCode": "77",
	"warehouseName": "FBA-25-EU"
}
"""


def init_data(rma_exampl,ignored_data):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        if data_pro is not None and data_pro == ignored_data:
            continue
        result[data_pro] = rma_exampl[data_pro]
    # print(result)
    return result


sku_str_list = '''
FTBFSD-0010FTBFSD-0010110
FTOFSF-0041FTOFSF-0041110
OEBO-0046OEBO-0046110
FRSSPF-0187FRSSPF-0187110
FTPLDB-2002FTPLDB-2002110
FTPLRG-0032FTPLRG-0032110


'''


filename = "D:/问题/sku-dif/EUUK库存拆分-1.xlsx"


# content = ''
# with open('tmpfile', 'r', encoding='utf-8') as file:
#     # content = file.read()  # 一次性读取整个文件内容
#     # 或者使用以下方式逐行读取
#     for line in file:
#         content += line
#     print(content)

# df = pd.read_excel(filename)
i = 2
# 加载现有的Excel文件
workbook = load_workbook(filename)
billNo = ''
# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    warehouseCode = row[1]
    skuCode = row[2]
    # oldSkuCode = row[3]
    storeName = row[5]
    platform = row[4]

    all_qty = row[6]
    gentime = datetime.now().strftime('%Y%m%d')
    billNo = gentime+ '_'+"_fix_dif_occpy"
    # if dif_qty > 0:
    #     continue
    invQty=0
    platform_code=""
    store_code=""
    print(skuCode,warehouseCode)
    stocklist = commonUtil.getInvstockbySku(skuCode,warehouseCode)
    stocklistN = []
    print(stocklist)
    for stock in stocklist:
        if storeName in stock["storeName"]:
            platform_code=stock["platform"]
            store_code=stock["store"]
            invQty+=stock["useQty"]
            stocklistN.append(stock)

    sheet.cell(row=i + 2, column=len(row) + 1, value=invQty)
    sheet.cell(row=i + 2, column=len(row) + 2, value=len(stocklistN))
    sheet.cell(row=i + 2, column=len(row) + 3, value=platform_code)
    sheet.cell(row=i + 2, column=len(row) + 4, value=store_code)
    req_exam = init_data(req_param,None)
    skulist =req_exam["skuList"]
    req_exam["warehouseCode"]=warehouseCode



    # print(type(update_time))

    req_exam["operationTypeEnum"] = "NULL"
    req_exam["remark"] = billNo
    req_exam["billNo"] = billNo
    req_exam["originBillNo"] = billNo

    # print(billTime)
    updateTime = datetime.now()
    # update_time = datetime.strftime(updateTime, "%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    # req_exam = json.dumps(req_exam)
    print(req_exam)
    #发送POST请求

    # response = requests.post(url, json=req_exam, headers=headers)
    # print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i>=1:
    #     break
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
workbook.save(f_tm + '_all_'  + '_fix_dif_stock.xlsx')
