import json
import time

from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

import pandas as pd

import commonUtil

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'
log_url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStockLog'
# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }


handle_order = '''
{
	"billNo": "GOC64758658",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-22T18:44:02.030",
	"billTypeEnum": "SALE_OUTBOUND_ORDERS",
	"operationTypeEnum": "HOLD",
	"originBillNo": "GOC64758658",
	"skuList": [
		{
			"badHoldQty": 0,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": 1,
			"inTransitQty": 0,
			"oldSkuCode": "OG0101A951",
			"platform": "10002",
			"skuCode": "TELC-0045",
			"store": "10334",
			"totalQty": 0,
			"useQty": -1
		}
	],
	"warehouseCode": "13",
	"warehouseName": "FBA-02"
}

'''

token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

order_status_complete = '''
GOC62772868
GOC62930611
GOC62969957
GOC62979830
GOC62979832
GOC62991917
GOC63010331
GOC63020057
GOC63041859
GOC63054674
GOC63054678
GOC63058799
GOC63058804
GOC63060673
GOC63072918
GOC63072919
GOC63074106
GOC63075688
GOC63077525
GOC63083345
GOC63084065
GOC63084094
GOC63085223
GOC63100422
GOC63134640
GOC63135425
GOC63135448
GOC63136508
GOC63138038
GOC63138386
GOC63138403
GOC63142372
GOC63146775
GOC63149074
GOC63158252
GOC63168398
GOC63168534
GOC63170073
GOC63173898
GOC63175170
GOC63176884
GOC63178068
GOC63179459
GOC63179839
GOC63181542
GOC63181781
GOC63184285
GOC63186847
GOC63187620
GOC63189285
GOC63190904
GOC63194381
GOC63196787
GOC63199981
GOC63203108
GOC63203695
GOC63203714
GOC63206095
GOC63206328
GOC63206493
GOC63207355
GOC63207753
GOC63207763
GOC63208059
GOC63208305
GOC63208600
GOC63209280
GOC63209718
GOC63209857
GOC63209900
GOC63213696
GOC63220451
GOC63222659
GOC63223064
GOC63223111
GOC63223269
GOC63223406
GOC63223451
GOC63223555
GOC63223572
GOC63223573
GOC63227339
GOC63229025
GOC63229228
GOC63231474
GOC63232184
GOC63233755
GOC63234095
GOC63235229
GOC63235883
GOC63237760
GOC63245525
GOC63247935
GOC63248312
GOC63249368
GOC63249563
GOC63249583
GOC63249643
GOC63249651
GOC63250229
GOC63250403
GOC63250562
GOC63250569
GOC63251314
GOC63252268
GOC63255487
GOC63255670
GOC63255680
GOC63255767
GOC63255794
GOC63255798
GOC63258654
GOC63258897
GOC63258913
GOC63259400
GOC63260011
GOC63260832
GOC63260959
GOC63262062
GOC63262444
GOC63262448
GOC63263072
GOC63264119
GOC63266045
GOC63269780
GOC63270527
GOC63270735
GOC63273480
GOC63273524
GOC63284090
GOC63284112
GOC63284127
GOC63284160
GOC63284176
GOC63284287
GOC63284309
GOC63284327
GOC63284329
GOC63284361
GOC63284389
GOC63284405
GOC63284445
GOC63284449
GOC63284462
GOC63284480
GOC63284483
GOC63284500
GOC63284501
GOC63284519
GOC63284552
GOC63284563
GOC63284583
GOC63284626
GOC63284641
GOC63284731
GOC63284791
GOC63284815
GOC63284850
GOC63284887
GOC63284895
GOC63284905
GOC63284952
GOC63284953
GOC63284965
GOC63284984
GOC63284990
GOC63284997
GOC63284998
GOC63285000
GOC63285040
GOC63285095
GOC63285148
GOC63285158
GOC63285161
GOC63285165
GOC63285179
GOC63285182
GOC63285227
GOC63286433
GOC63286581
GOC63286583
GOC63287635
GOC63289642
GOC63289650
GOC63290432
GOC63291095
GOC63293428
GOC63294054
GOC63294096
GOC63294291
GOC63294433
GOC63296384
GOC63296777
GOC63297234
GOC63297515
GOC63297606
GOC63297679
GOC63297682
GOC63298242
GOC63299400
GOC63299403
GOC63299405
GOC63299407
GOC63299410
GOC63299411
GOC63300100
GOC63300595
GOC63301150
GOC63302664
GOC63302781
GOC63302813
GOC63303452
GOC63303511
GOC63304637
GOC63304778
GOC63305302
GOC63306650
GOC63307892
GOC63307893
GOC63308103
GOC63308789
GOC63309322
GOC63309553
GOC63310121
GOC63310820
GOC63311237
GOC63311751
GOC63311812
GOC63311814
GOC63311815
GOC63311816
GOC63311817
GOC63311819
GOC63311822
GOC63311826
GOC63311827
GOC63311828
GOC63311831
GOC63311832
GOC63311833
GOC63311834
GOC63311838
GOC63311839
GOC63311841
GOC63311844
GOC63311848
GOC63311852
GOC63311853
GOC63311857
GOC63311860
GOC63311863
GOC63311867
GOC63311869
GOC63311873
GOC63311875
GOC63311877
GOC63311879
GOC63311880
GOC63311883
GOC63311887
GOC63311888
GOC63311889
GOC63311893
GOC63311894
GOC63311898
GOC63311902
GOC63311905
GOC63311908
GOC63311912
GOC63311916
GOC63311919
GOC63311922
GOC63311925
GOC63311926
GOC63311930
GOC63311933
GOC63311934
GOC63311936
GOC63311937
GOC63311939
GOC63311943
GOC63311947
GOC63312067
GOC63312572
GOC63314017
GOC63314453
GOC63314545
GOC63314548
GOC63314721
GOC63316278
GOC63316694
GOC63317186
GOC63318607
GOC63319578
GOC63321822
GOC63323199
GOC63323456
GOC63323460
GOC63323634
GOC63323758
GOC63323812
GOC63323859
GOC63323954
GOC63323958
GOC63324189
GOC63324248
GOC63324406
GOC63324416
GOC63324493
GOC63324544
GOC63324869
GOC63324955
GOC63324996
GOC63325001
GOC63325005
GOC63325275
GOC63325424
GOC63325592
GOC63325629
GOC63325779
GOC63325786
GOC63326571
GOC63326575
GOC63326840
GOC63326945
GOC63326963
GOC63327301
GOC63327349
GOC63327351
GOC63327489
GOC63328202
GOC63328205
GOC63328410
GOC63328726
GOC63328732
GOC63328928
GOC63329026
GOC63329134
GOC63329182
GOC63329288
GOC63329399
GOC63329735
GOC63329904
GOC63330708
GOC63330944
GOC63331133
GOC63331183
GOC63331186
GOC63331327
GOC63331535
GOC63331703
GOC63331980
GOC63332021
GOC63332023
GOC63332049
GOC63332079
GOC63332131
GOC63332205
GOC63332363
GOC63332365
GOC63332577
GOC63332649
GOC63332987
GOC63333170
GOC63333176
GOC63333779
GOC63334021
GOC63334024
GOC63334116
GOC63334286
GOC63334287
GOC63334415
GOC63334497
GOC63334719
GOC63334947
GOC63335010
GOC63335078
GOC63335173
GOC63335259
GOC63335570
GOC63335798
GOC63335813
GOC63336231
GOC63336335
GOC63336425
GOC63336538
GOC63336600
GOC63336635
GOC63337022
GOC63337163
GOC63337164
GOC63337271
GOC63337289
GOC63337479
GOC63337483
GOC63337859
GOC63337986
GOC63337993
GOC63338039
GOC63338043
GOC63338047
GOC63338049
GOC63338051
GOC63338211
GOC63338339
GOC63338343
GOC63338449
GOC63338609
GOC63338849
GOC63338912
GOC63338977
GOC63339162
GOC63339168
GOC63339564
GOC63339683
GOC63339687
GOC63340285
GOC63340492
GOC63340508
GOC63340542
GOC63340716
GOC63340751
GOC63341071
GOC63341083
GOC63341090
GOC63341097
GOC63341100
GOC63341102
GOC63341434
GOC63341582
GOC63341585
GOC63341586
GOC63341589
GOC63341593
GOC63341605
GOC63341719
GOC63341723
GOC63342085
GOC63342112
GOC63342130
GOC63342214
GOC63342216
GOC63342258
GOC63342275
GOC63342332
GOC63342339
GOC63342538
GOC63342675
GOC63342693
GOC63342818
GOC63343004
GOC63343015
GOC63343070
GOC63343130
GOC63343189
GOC63343312
GOC63343439
GOC63343445
GOC63343721
GOC63343941
GOC63344301
GOC63344522
GOC63344780
GOC63345139
GOC63345279
GOC63345406
GOC63346265
GOC63346375
GOC63346431
GOC63346680
GOC63346780
GOC63346785
GOC63346952
GOC63347035
GOC63347220
GOC63347570
GOC63347583
GOC63347790
GOC63347793
GOC63348607
GOC63349188
GOC63349242
GOC63349593
GOC63349620
GOC63349749
GOC63349752
GOC63349805
GOC63350000
GOC63350158
GOC63350168
GOC63350341
GOC63350775
GOC63350922
GOC63351282
GOC63351739
GOC63351841
GOC63351966
GOC63352462
GOC63352796
GOC63352808
GOC63352841
GOC63352974
GOC63352983
GOC63353074
GOC63353076
GOC63353080
GOC63353251
GOC63353420
GOC63353462
GOC63353533
GOC63353626
GOC63353627
GOC63353635
GOC63353636
GOC63353650
GOC63353653
GOC63353881
GOC63354254
GOC63354736
GOC63355853
GOC63357272
GOC63357274
GOC63358160
GOC63359229
GOC63359625
GOC63359778
GOC63359784
GOC63359789
GOC63359792
GOC63359794
GOC63359796
GOC63359800
GOC63360218
GOC63362712
GOC63363468
GOC63363471
GOC63363474
GOC63363834
GOC63364430
GOC63365970
GOC63366262
GOC63366285
GOC63366287
GOC63366296
GOC63366303
GOC63366308
GOC63366900
GOC63367114
GOC63367192
GOC63367212
GOC63367616
GOC63368460
GOC63368464
GOC63369143
GOC63369369
GOC63369492
GOC63370391
GOC63370393
GOC63370400
GOC63370779
GOC63370875
GOC63371137
GOC63372579
GOC63373206
GOC63373625
GOC63374406
GOC63376130
GOC63376792
GOC63376908
GOC63376909
GOC63376911
GOC63376913
GOC63377003
GOC63377005
GOC63378163
GOC63378164
GOC63378289
GOC63378570
GOC63378574
GOC63378575
GOC63378577
GOC63378925
GOC63380792
GOC63382755
GOC63383369
GOC63384083
GOC63384795
GOC63385633
GOC63385728
GOC63385988
GOC63387465
GOC63387730
GOC63389624
GOC63390187
GOC63391333
GOC63393129
GOC63394271
GOC63394274
GOC63398255
GOC63405960
GOC63406224
GOC63412975
GOC63430470
GOC63455592
GOC63460444
GOC63460448
GOC63460648
GOC63460783
GOC63466794
GOC63488656
GOC63498917
GOC63499240
GOC63502217
GOC63502619
GOC63502840
GOC63517019
GOC63517129
GOC63530931
GOC63534636
GOC63558037
GOC63561435
GOC63582289
GOC63584954
GOC63586145
GOC63588090
GOC63591338
GOC63592925
GOC63594614
GOC63597737
GOC63599121
GOC63601558
GOC63605387
GOC63606648
GOC63607643
GOC63608692
GOC63613575
GOC63613576
GOC63622761
GOC63624212
GOC63624297
GOC63628548
GOC63632259
GOC63648762
GOC63648764
GOC63649613
GOC63651429
GOC63655191
GOC63661935
GOC63674347
GOC63678688
GOC63694308
GOC63696190
GOC63705514
GOC63718020
GOC63719192
GOC63722217
GOC63731528
GOC63741471
GOC63762221
GOC63763269
GOC63763777
GOC63764950
GOC63781699
GOC63790441
GOC63793244
GOC63810078
GOC63831826
GOC63836026
GOC63884258
GOC63902111
GOC63904455
GOC63915594
GOC63920446
GOC63930802
GOC63974276
GOC64009812
GOC64029769
GOC64063330
GOC64069443
GOC64106667
GOC64127770
GOC64136645
GOC64150030
GOC64303634
GOC64497461





'''

order_duplicate="""





"""

order_exist = """

"""

def init_data(rma_exampl):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        result[data_pro] = rma_exampl[data_pro]
        if data_pro == "skuList":
            for sku in rma_exampl[data_pro]:
                for sku_pro in sku:
                    result[data_pro][sku_pro] = sku[sku_pro]
    return result

use_qty_str = ''
with open('use_qty', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str += line
    # print(use_qty_str)

use_qty_str = json.loads(use_qty_str)



occpy_qty = ''
with open('occpy_qty', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        occpy_qty += line
    # print(use_qty_str)

occpy_qty = json.loads(occpy_qty)

use_qty_str_log = ''
with open('use_qty_log', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str_log += line
    # print(use_qty_str)

use_qty_str_log = json.loads(use_qty_str_log)

filename = "d:\\tmp\\执行结果0527_sale_error.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)


# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook =load_workbook(filename)

# 选择工作表
sheet = workbook.active
i=0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    skuCode = row[9]
    oldSkuCode = row[10]
    platform = row[7]
    store = row[8]
    billNo = row[3]
    origin_bill_no = row[2]
    bill_time = row[15]
    warehouse_code = row[6]
    cur_qty = int(row[11])
    update_time = row[15]
    req_exam = commonUtil.init_data(handle_order,None)
    req_exam["billNo"] = billNo
    req_exam["remark"] = billNo
    req_exam["originBillNo"] = billNo
    # print(type(update_time))
    # print(update_time)
    updateTime = datetime.fromisoformat(update_time)
    # update_time = datetime.strftime(update_time,"%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    req_exam["warehouseCode"] = warehouse_code
    row_sku_str=skuCode+oldSkuCode+platform+store+warehouse_code
    skuList = req_exam["skuList"]
    for sku in skuList:
        sku["platform"] = platform
        sku["store"] = store
        sku["skuCode"] = skuCode
        sku["oldSkuCode"] = oldSkuCode
        sku["holdQty"] = cur_qty*-1
        sku["useQty"] = 0
        sku["totalQty"] = cur_qty*-1
    # req_exam = json.dumps(req_exam)
    # print(req_exam)

    if billNo in order_status_complete:
        if row_sku_str not in use_qty_str_log or use_qty_str_log[row_sku_str]-cur_qty<0:
            req_exam["skuList"][0]["holdQty"] = cur_qty
            req_exam["skuList"][0]["totalQty"] = cur_qty
            req_exam["operationTypeEnum"]="NULL"
            # req_exam = json.dumps(req_exam)
            print(req_exam)
            # 发送POST请求
            response = requests.post(log_url, json=req_exam, headers=headers)
            print(response.text)
            # req_exam = json.loads(req_exam)
            req_exam["skuList"][0]["holdQty"] = cur_qty*-1
            req_exam["skuList"][0]["totalQty"] = cur_qty*-1
            req_exam["operationTypeEnum"] = "HOLD"
            # req_exam = json.dumps(req_exam)
    else:
        continue

    # if bill_no in order_status_normal:

    # req_exam = json.dumps(req_exam)
    print(req_exam)

    time.sleep(0.5)
    #    发送POST请求
    response = requests.post(url, json=req_exam, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i +=1
    # if i>=1:
    #     break
print(i)
use_qty_str_log = json.dumps(use_qty_str_log)
print(use_qty_str_log)
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
workbook.save(ts+"_"+"tmp-sale_orders.xlsx")