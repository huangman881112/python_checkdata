from datetime import datetime
import secrets
import string
import time

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)

sku_str_List = '''
FTLFSO-1002-EFTLFSO-1002-E65
FTPLWP-0020-EFTPLWP-0020-E78
FTPLWP-0017-EFTPLWP-0017-E78
FTLFSO-1003-EFTLFSO-1003-E65
FTPLWP-0021-EFTPLWP-0021-E78
FTPLPU-0022-EFTPLPU-0022-E78
FTPLWP-0016-EFTPLWP-0016-E78
FTPLWP-0023-EFTPLWP-0023-E78
FTLFSO-1001-EFTLFSO-1001-E65
FTPLPU-0023-EFTPLPU-0023-E78
FTPLWP-0015-EFTPLWP-0015-E78
FTPLPU-0021-EFTPLPU-0021-E78
FTPLCG-0564-EFTPLCG-0564-E78
FTPLWP-0014-EFTPLWP-0014-E78

'''

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

head_out = """
{
	"operationTypeEnum": "HEAD_IN_TRANSIT",
	"skuList": [
		{
			"freezeQty": 0,
			"holdQty": 0,
			"useQty": 1,
			"totalQty": 1,
			"inTransitQty": -1,
			"store": "11104",
			"platform": "10002",
			"badQty": 0,
			"oldSkuCode": "FTBFSD-0026-E",
			"skuCode": "FTBFSD-0026-E",
			"badHoldQty": 0
		}
	],
	"originBillNo": "2401H-S2-FDE",
	"warehouseName": "FBA-22-EU",
	"billTypeEnum": "FBA_INBOUND",
	"warehouseCode": "65",
	"billTime": 1716450622413,
	"billNo": "DB24052300557",
	"billStatusEnum": "FINISH"
}
"""


def init_data(rma_exampl):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        result[data_pro] = rma_exampl[data_pro]
    # print(result)
    return result


def geneare_str(length):
    characters = string.ascii_letters + string.digits
    random_string = ''.join(secrets.choice(characters) for _ in range(length))
    return random_string

filename = "d:\\tmp\\sku_dif\\20240523_204302_2401H-S18-FDE_list_json2excel.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)
billNo = ''
# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    skuCode = row[1]
    oldSkuCode = row[9]
    platform = row[12]
    store = row[13]
    warehouse_code = row[2]
    row_sku_str = skuCode+oldSkuCode+warehouse_code
    # print(row_sku_str)
    if row_sku_str not in sku_str_List:
        continue

    billNo = row[0]
    origin_bill_no = row[0]
    bill_time = row[8]
    up_qty = int(row[11])
    if up_qty == 0 :
        continue
    # update_time = row[15]
    req_exam = init_data(head_out)
    gentime = datetime.now().strftime('%m%d%H')
    if billNo == '':
        billNo = gentime + geneare_str(4) + "_invfix"
    req_exam["billNo"] = billNo
    req_exam["remark"] = billNo
    req_exam["originBillNo"] = billNo
    req_exam["operationTypeEnum"]="NULL"
    # print(type(update_time))
    print(bill_time)
    updateTime = datetime.fromisoformat(bill_time)
    # update_time = datetime.strftime(update_time,"%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    req_exam["warehouseCode"] = warehouse_code
    skuList = req_exam["skuList"]
    for sku in skuList:
        sku["platform"] = platform
        sku["store"] = store
        sku["skuCode"] = skuCode
        sku["oldSkuCode"] = oldSkuCode
        sku["inTransitQty"] = 0
        sku["useQty"] = up_qty*-1
        sku["totalQty"] = up_qty*-1

    # req_exam = json.dumps(req_exam)
    print(req_exam)
    # 发送POST请求
    response = requests.post(url, json=req_exam, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i >=1:
    #     break
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
# workbook.save(f_tm +'_' +billNo+'_head_调出.xlsx')
