from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    # print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

head_inbound_exampl = """
{
	"billNo": "PAP20240415084",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-10T07:01:26.199",
	"billTypeEnum": "OTHER_OUTBOUND",
	"operationTypeEnum": "PUTAWAY",
	"originBillNo": "2403H-S191-CA03",
	"skuList": [
		{
			"badHoldQty": 0,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": 533,
			"inTransitQty": -533,
			"oldSkuCode": "FTPLDB-0005",
			"platform": "10001",
			"skuCode": "FTPLDB-0005",
			"store": "10415",
			"totalQty": 533,
			"useQty": 0
		}
	],
	"warehouseCode": "112"
}

"""


def init_data(rma_exampl):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        result[data_pro] = rma_exampl[data_pro]
    # print(result)
    return result


filename = "d:\\tmp\\tmp-head_入库.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)

# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break

    new_value = row[14]
    load_value = json.loads(new_value)
    load_value["req"]["remark"] = load_value["req"]["billNo"]
    # print(load_value["req"]["billNo"])
    billno = load_value["req"]["billNo"]
    sheet.cell(i + 2, len(row) + 1, load_value["req"]["originBillNo"])
    tm = load_value["req"]["billTime"]
    tm = time.localtime(tm / 1000)
    tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
    load_value["req"]["billTime"] = tm
    load_value["req"]["operationTypeEnum"]="PUTAWAY"
    load_value = load_value["req"]
    skuList=load_value["skuList"]
    for sku in skuList:
        sku["inTransitQty"] = 0

    # load_value = json.dumps(load_value)
    print(load_value)
    # 发送POST请求
    response = requests.post(url, json=load_value, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i>=1:
    #     break
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
# workbook.save(f_tm + 'example.xlsx')
""
