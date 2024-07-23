from datetime import datetime
import secrets
import string
import time

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import requests
import json

import commonUtil as cul




zpoNo = ""
# skuCode = "FTPLPB-0292"
bill_nolist = []
skuList  =[]
sku_str= "FTLFTS-6043"
# skuList.append(skuCode)
bill_nolist.append(zpoNo)

skuList = sku_str.split(",",-1)
res={}
zpo_res_list = []
zpo_res_list_ = []


def generateExec(hears,data):
    # 加载现有的Excel文件
    workbook = Workbook()
    # 选择工作表
    sheet = workbook.active
    sheet.cell(1, 1, "skucode")
    sheet.cell(1, 2, "oldSku")
    sheet.cell(1, 3, "zpo")
    sheet.cell(1, 4, "receiveQty")
    sheet.cell(1, 5, "actualQty")
    sheet.cell(1, 6, "po")
    file_time = cul.FILE_PROFIX + datetime.now().strftime('%Y%m%d_%H%M%S')
    workbook.save(file_time + "_" + "inv" + "_" + '采购收货在途.xlsx')
def getZpoRes(zpos, planNolist):

    i=1
    for planno in planNolist:

        if str(planno).startswith("2"):
            param_body = cul.getApiList()[cul.HEAD_SUPPLYSTRAIGHT_DETAIL_PARA_BODY]
            param_body[cul.HEAD_SUPPLYSTRAIGHT_DETAIL_PARA_BODY_PLANNO] = planno
            # param_body[comutils.PURCHASE_ZPO_LIST_PARA_BODY_SKUCODE] = sku
            response = requests.post(cul.getApiList()[cul.HEAD_SUPPLYSTRAIGHT_DETAIL_URL], json=param_body, headers=cul.getHeaders())
            # print(response.text)
            po_res = json.loads(response.text)
            zpo_list = po_res[cul.DATA]
            # print(po_list)
            # res["zpo_inv_qty"] = 0
            sum_purchase_amount = 0
            for po_no in zpo_list:
                res = dict()
                if po_no["zPo"] != zpos:
                    continue
                i+=1
                res["cinvCode"] = po_no["cinvCode"]
                res["oldSku"] = po_no["oldSku"]
                res["zPo"] = po_no["zPo"]
                res["receiveQty"] = po_no["receiveQty"]
                res["actualQty"] = po_no["actualQty"]
                res["planno"] = planno
                zpo_res_list.append(res)

    # file_time = cul.FILE_PROFIX + datetime.now().strftime('%Y%m%d_%H%M%S')
    # workbook.save(file_time + "_" + zpos + "_" + '采购收货在途.xlsx')
    print(zpo_res_list)
    # print(zpo_res_list_)

# ZPO23121300310,ZPO24031900103
billnos = cul.getAllbilllist("FTBFSD-0106",[0,1,4],[])
print(billnos)
getZpoRes("ZPO24040900075",billnos)