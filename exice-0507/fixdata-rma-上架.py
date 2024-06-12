import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": "bef70aec-2ae2-46a7-b110-6871667c0bd7"  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

rma_exampl = """
{
	"operationTypeEnum": "PUTAWAY",
	"skuList": [
		{
			"freezeQty": 0,
			"holdQty": 0,
			"useQty": 0,
			"totalQty": 0,
			"inTransitQty": 0,
			"store": "10415",
			"platform": "10001",
			"badQty": 1,
			"oldSkuCode": "HEFTCF-6016",
			"skuCode": "HEFTCF-6016",
			"badHoldQty": -1
		}
	],
	"originBillNo": "QCO2024040500151",
	"warehouseName": "NJ正品仓",
	"cooperator": "赢他",
	"billTypeEnum": "RMA_RETURN_PUTAWA",
	"warehouseCode": "61",
	"billTime": 1714496912241,
	"whType": "oneself",
	"billNo": "PAP20240405245",
	"billStatusEnum": "FINISH"
}

"""


def init_data(rma_exampl):
    result = {}
    for data_pro in rma_exampl:
        result[data_pro] = rma_exampl[data_pro]
        if data_pro == "skuList":
            for sku in rma_exampl[data_pro]:
                for sku_pro in sku:
                    result[data_pro][sku_pro] = sku[sku_pro]
    return result


filename = "d:\\tmp\\tmp-rma.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)

# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    new_value = row[14]
    load_value = json.loads(new_value)
    load_value["req"]["remark"] = load_value["req"]["billNo"]
    tm = load_value["req"]["billTime"]
    print(tm)
    tm = time.localtime(tm / 1000)
    tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
    load_value["req"]["billTime"] = tm
    load_value["req"]["operationTypeEnum"] = "NULL"
    sku_list = load_value["req"]["skuList"]
    for sku in sku_list:
        sku["badHoldQty"] = 0

    load_value = json.dumps(load_value["req"])
    print(load_value)
    # 发送POST请求
    # response = requests.post(url, json=load_value, headers=headers)
    # print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
f_tm = datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
workbook.save(f_tm + 'example.xlsx')
