import json
import time

from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

import pandas as pd
import commonUtil as comutil
import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }



# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": comutil.getToken() # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}






order_status_complete = '''
GOC64869137
GOC64869186
GOC64869381
GOC64869485
GOC64869524
GOC64869560
GOC64869586
GOC64869608
GOC64869980
GOC64870008
GOC64870201
GOC64874677
GOC64875204
GOC64886315
GOC64895063
GOC64895081
GOC64900611
GOC64911382
GOC64869338
GOC64869391
GOC64870018
GOC64870155
GOC64870169
GOC64877207
GOC64881563
GOC64886164
GOC64895219
GOC64901024
GOC64901032
GOC64906821
GOC64908521
GOC64910561
GOC64914722
GOC64915798





'''

order_duplicate="""
GOC64900972



"""

order_exist = """

"""



use_qty_str = ''
with open('use_qty', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str += line
    # print(use_qty_str)

use_qty_str = json.loads(use_qty_str)

use_qty_str_log = ''
with open('use_qty_log', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str_log += line
    # print(use_qty_str)

use_qty_str_log = json.loads(use_qty_str_log)

filename = "d:\\tmp\\tmp.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"




# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook =load_workbook(filename)
apilist = comutil.getApiList()
# 选择工作表
sheet = workbook.active
i=0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):

    if row[0] is None:
        break
    new_value = row[14]
    load_value = json.loads(new_value)
    load_value["req"]["remark"] = load_value["req"]["billNo"]
    bill_no = load_value["req"]["billNo"]
    sheet.cell(row=i+2, column=15).value = json.dumps(load_value)
    tm = load_value["req"]["billTime"]
    tm = time.localtime(tm/1000)
    row_sku_str=''
    tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
    load_value["req"]["billTime"]= tm
    load_value = load_value["req"]
    sku_list = load_value["skuList"]
    warehouse_code = load_value["warehouseCode"]
    cur_qty = 0
    if len(sku_list)>1:
        sheet.cell(row=i+2, column=len(row)+1, value="多sku")
    else:
        for sku in sku_list:
            sheet.cell(row=i + 2, column=len(row)+1, value="单sku")
            sheet.cell(row=i + 2, column=len(row)+2, value=sku["platform"])
            sheet.cell(row=i + 2, column=len(row)+3, value=sku["store"])
            cur_qty=sku["holdQty"]
            sheet.cell(row=i + 2, column=len(row)+4, value=sku["useQty"])
            row_sku_str = sku["skuCode"]+sku["oldSkuCode"]+sku["platform"]+sku["store"]+warehouse_code

    # if bill_no not in order_duplicate and bill_no not in order_exist:
    #     print(bill_no)
    #     if row_sku_str in use_qty_str:
    #         if use_qty_str[row_sku_str]-cur_qty>=0 and use_qty_str_log[row_sku_str]-cur_qty>=0 and bill_no in order_status_complete :
    #             # print(use_qty_str[row_sku_str])
    #             # print(cur_qty)
    #             use_qty_str[row_sku_str] -=cur_qty
    #             use_qty_str_log[row_sku_str] -=cur_qty
    #             # print(use_qty_str[row_sku_str])
    #             print(row_sku_str,"_",bill_no)
    #         else:
    #             continue
    #     else:
    #         continue
    # else:
    #     continue

    # /., / if bill_no in order_status_normal:

    # load_value = json.dumps(load_value)
    print(load_value)
    # print(apilist[comutil.STOCK_URL])
    # 发送POST请求
    # response = requests.post(apilist[comutil.STOCK_URL], json=load_value, headers=headers)
    # print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i +=1
    # if i>=1:
    #     break
print(i)
use_qty_str = json.dumps(use_qty_str)
print(use_qty_str)
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
workbook.save(ts+"_"+"tmp-sale_orders.xlsx")