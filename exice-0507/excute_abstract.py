import Action
from datetime import datetime
import json
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import json
import commonUtil as comutils

import Action





def excute_error_log(filename, action: Action, m, flag,wait_time=0):
    workbook = load_workbook(filename)
    po_zip_list = {}
    # 选择工作表
    i = 0
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        id =row[0]
        billType = row[1]
        new_value = row[14]
        skuCode = row[4]
        print(new_value)
        load_value = json.loads(new_value)
        load_value["req"]["remark"] = load_value["req"]["billNo"]
        # print(load_value["req"]["billNo"])
        billno = load_value["req"]["billNo"]
        # sheet.cell(i + 2, len(row) + 1, load_value["req"]["originBillNo"])
        tm = load_value["req"]["billTime"]
        tm = time.localtime(tm / 1000)
        tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
        load_value["req"]["billTime"] = tm
        # load_value["req"]["operationTypeEnum"]="NULL"
        load_value = load_value["req"]

        action_new = action.getActionByType(billType)
        if action_new is None:
            continue
        errorLogIdList=[]
        errorLogIdList.append(id)
        load_value["errorLogIdList"] = errorLogIdList
        load_value["repairOpt"] = True
        load_value = action_new.excute_req(load_value,skuCode)
        if flag == False:
            load_value = json.dumps(load_value)
            print(load_value)
        if flag == True and load_value is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_URL], json=load_value,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        time.sleep(wait_time)
        if m > 0 and i >= m:
            break
    return i

action = {}


def excute_occpy(filename, req_exam_str, action: Action, m, flag):
    i = 0
    # 加载现有的Excel文件
    workbook = load_workbook(filename)
    billNo = ''
    # 选择工作表
    idlist = []
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        sku_exam = {}
        id = row[0]
        biz_type = row[1]
        skuCode = row[10]
        oldSkuCode = row[11]
        platform = row[8]
        store = row[9]
        site = row[18]
        sku_exam["skuCode"] = skuCode
        sku_exam["oldSkuCode"] = oldSkuCode
        sku_exam["platform"] = platform
        sku_exam["store"] = store
        sku_exam["site"] = site
        if row[19] != "":
            billNo = row[20]
            origin_bill_no = row[19]
        else:
            billNo = row[4]
            origin_bill_no = row[4]
        bill_time = row[16]
        warehouse_code = row[7]
        qty = int(row[12])
        if row[21]!=None:
            qty= int(row[21])
        update_time = bill_time
        if req_exam_str is None:
            req_exam_str = comutils.OTHER_OUTBOUND
        req_exam = comutils.init_data(req_exam_str, None)
        req_exam["billNo"] = billNo
        if biz_type in comutils.getOperType():
            req_exam["operationTypeEnum"]=comutils.getOperType()[biz_type]
        req_exam["remark"] = billNo
        req_exam["originBillNo"] = origin_bill_no
        # print(skuCode + oldSkuCode + platform + store + warehouse_code)
        updateTime = datetime.fromisoformat(update_time)
        req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
        req_exam["warehouseCode"] = warehouse_code
        idlist.append(id)
        req_exam["errorLogIdList"] = idlist
        req_exam["repairOpt"] = True
        skuList = req_exam["skuList"]
        for sku in skuList:
            for prop in sku:
                if sku[prop] == 0 or sku[prop] == '':
                    continue
                elif prop.endswith('Qty'):
                    sku[prop] = qty
                else:
                    sku[prop] = sku_exam[prop]
            if site is None:
                skuList.remove(sku)
                sku = comutils.init_data(sku, "site", "json")
                skuList.append(sku)
        # 发送POST请求
        # req_exam["skuList"] = action.excute_sku(skuList)
        req_exam = action.excute_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
            print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break
    return i

def execute_skudif(filename, action: Action, m, flag):
    workbook = load_workbook(filename)
    billNo = ''
    i = 0
    # 选择工作表
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        sku_exam = {}
        skuCode = row[7]
        oldSkuCode = row[8]
        warehouseCode = row[10]
        warehouseName = str(row[9])
        site = str(row[11])
        # print(warehouseName)
        if warehouseName.startswith("FBA") or warehouseName.startswith("CG") or warehouseName.startswith("WFS"):
            continue
        # platform = row[4]
        # store = row[5]
        dif_qty = int(row[11])
        sku_exam["skuCode"] = skuCode
        sku_exam["oldSkuCode"] = oldSkuCode
        sku_exam["site"] = site
        gentime = datetime.now().strftime('%Y%m%d')
        billNo = gentime + "_fix_dif_stock"
        if dif_qty > 0:
            head_trans_out = comutils.OTHER_OUTBOUND
        else:
            head_trans_out = comutils.OTHER_INBOUND

        req_exam = comutils.init_data(head_trans_out, None)

        req_exam["warehouseCode"] = warehouseCode
        req_exam["billNo"] = billNo
        req_exam["originBillNo"] = billNo
        req_exam["remark"] = billNo
        skustr = skuCode + oldSkuCode + warehouseCode
        # print(type(update_time))
        skuList = req_exam["skuList"]
        for sku in skuList:
            for prop in sku:
                if sku[prop] == 0 or sku[prop] == '':
                    continue
                elif prop.endswith('Qty'):
                    sku[prop] = dif_qty
                elif prop in sku_exam:
                    sku[prop] = sku_exam[prop]
            if site == "None":
                skuList.remove(sku)
                sku = comutils.init_data(sku,"site","json")
                skuList.append(sku)
        # print(billTime)
        # 发送POST请求
        req_exam = action.execute_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
            print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break
    return i


def execute_stock_with_unit(filename, req_exam_str, action: Action, m, flag):
    req_str=""
    workbook = load_workbook(filename)
    billNo = ''
    i = 0
    # 选择工作表
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        sku_exam = {}
        skuCode = row[0]
        oldSkuCode = row[1]
        platform = row[2]
        store = row[3]
        warehouseCode = str(row[4])
        site = str(row[5])
        warehouseName = comutils.getwarehouse_list()[warehouseCode]
        # print(warehouseName)
        # if warehouseName.startswith("FBA") or warehouseName.startswith("CG") or warehouseName.startswith("WFS"):
        #     continue
        dif_qty = int(row[6])
        if len(row) >7:
            billNo = row[7]
        sku_exam["skuCode"] = skuCode
        sku_exam["oldSkuCode"] = oldSkuCode
        sku_exam["platform"] = platform
        sku_exam["store"] = store
        if site != "":
            sku_exam["site"] = site
        gentime = datetime.now().strftime('%Y%m%d')
        if billNo is None or billNo == '':
            billNo = gentime + "_fix_dif_site_stock"
        if dif_qty == 0:
            continue
        if req_exam_str is None:
            if dif_qty > 0:
                req_str = comutils.OTHER_INBOUND
            else:
                req_str = comutils.OTHER_OUTBOUND
        else:
            req_str = req_exam_str
        req_exam = comutils.init_data(req_str, None)
        req_exam["warehouseCode"] = warehouseCode
        req_exam["billNo"] = billNo
        req_exam["originBillNo"] = billNo
        req_exam["remark"] = billNo
        req_exam["repairOpt"] = True
        req_exam["billTime"] = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        skustr = skuCode + oldSkuCode + warehouseCode
        # print(type(update_time))
        skuList = req_exam["skuList"]
        for sku in skuList:
            for prop in sku:
                if sku[prop] == 0 or sku[prop] == '':
                    continue
                elif prop.endswith('Qty'):
                    sku[prop] = dif_qty
                elif prop in sku_exam:
                    sku[prop] = sku_exam[prop]
            if site == "None":
                skuList.remove(sku)
                sku = comutils.init_data(sku,"site","json")
                skuList.append(sku)
        # print(billTime)
        # 发送POST请求
        req_exam = action.execute_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
            print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break
    return i

def execute_send_log(filename, req_exam_str, action: Action, m, flag):
    workbook = load_workbook(filename)
    # 选择工作表
    sheet = workbook.active
    i=0;
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        message_value = row[2]
        req_exam = json.loads(message_value)
        for data in req_exam:
            data["remark"] = data["billNo"]
        req_exam = action.execute_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
            print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_LIST_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break
    return i

def excute_sku(action: Action):
    action.handle_sku()


def excute_req(action: Action):
    action.handle_req()
