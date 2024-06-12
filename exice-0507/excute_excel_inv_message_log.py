import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": "1fe24d9f-294a-4fcd-8e0f-fb850a38235e" # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

filename = "d:\\tmp\\2024-05-08_18_10_35excute_excel_inv_message_log.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)


# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)

# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    message_value = row[2]
    load_value = json.loads(message_value)
    for data in load_value:
        sheet.cell(i+2, len(row)+1, data["billTypeEnum"])
        sheet.cell(i+2, len(row)+2, data["billNo"])

        if data["billNo"] in content:
            sheet.cell(i+2, len(row)+3, "已执行")
        else:
            sheet.cell(i+2, len(row)+3, "未执行")
        sheet.cell(i + 2, len(row) + 4, data["skuList"][0]["skuCode"])
        sheet.cell(i + 2, len(row) + 5, data["skuList"][0]["oldSkuCode"])
        sheet.cell(i + 2, len(row) + 6, data["warehouseCode"])

    # load_value["req"]["remark"] = load_value["req"]["billNo"]
    # print(load_value["req"]["billNo"])
    # sheet.cell(i+2, len(row)+1, load_value["req"]["billNo"])
    # tm = load_value["req"]["billTime"]
    # print(tm)
    # tm = time.localtime(tm/1000)
    #
    # tm = time.strftime('%Y-%m-%d %H:%M:%S', tm)
    # load_value["req"]["billTime"]= tm
    # print(load_value)
    # load_value = json.dumps(load_value)
    # 发送POST请求
    # response = requests.post(url, json=load_value, headers=headers)
    # print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
f_tm= datetime.datetime.now().strftime('%Y-%m-%d_%H_%M_%S')
print(i)
workbook.save(f_tm + 'excute_excel_inv_message_log.xlsx')