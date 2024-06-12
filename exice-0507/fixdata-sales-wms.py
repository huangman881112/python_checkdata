import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }

token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

use_qty_str = ''
with open('use_qty', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str += line
    # print(use_qty_str)

use_qty_str = json.loads(use_qty_str)

use_qty_str_log = ''
with open('use_qty_log', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str_log += line
    # print(use_qty_str)

use_qty_str_log = json.loads(use_qty_str_log)



sales_order_exist = '''
GOC64992817
GOC65005601
'''

bill_no_exist ="""
ODO202405261422
ODO202405264122
ODO202405275270
ODO202405266037
ODO20240528916
"""


filename = "d:\\tmp\\tmp-wms.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)


# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)
j=0
# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    i += 1
    if row[0] is None:
        break
    new_value = row[14]
    load_value = json.loads(new_value)
    load_value["req"]["remark"] = load_value["req"]["billNo"]
    billNo = load_value["req"]["billNo"]
    tm = load_value["req"]["billTime"]
    origin_bill_no = load_value["req"]["originBillNo"]

    tm = time.localtime(tm/1000)
    sheet.cell(row=i+1, column=len(row)+1, value=origin_bill_no)
    tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
    load_value["req"]["billTime"]= tm
    sku_list= load_value["req"]["skuList"]
    warehouse_code = load_value["req"]["warehouseCode"]
    sku = sku_list[0]
    # if sku["holdQty"] == 0:
    #     sku["holdQty"]= sku["useQty"]
    #     sku["useQty"]= 0
    sheet.cell(row=i + 1, column=len(row) + 2, value=sku["platform"])
    sheet.cell(row=i + 1, column=len(row) + 3, value=sku["store"])
    sheet.cell(row=i + 1, column=len(row) + 4, value=sku["holdQty"])
    row_sku_str = sku["skuCode"]+sku["oldSkuCode"]+sku["platform"]+sku["store"]+warehouse_code
    sheet.cell(row=i + 1, column=len(row) + 5, value=row_sku_str)

    # if origin_bill_no in sales_order_exist:
    #     use_qty_str_log[row_sku_str] += sku["holdQty"]
    #     print(origin_bill_no)
    # elif (origin_bill_no not in sales_order_exist and row_sku_str in use_qty_str
    #       and billNo not in bill_no_exist
    #       and use_qty_str[row_sku_str]+sku["holdQty"] >= 0
    #       and use_qty_str_log[row_sku_str]+sku["holdQty"] >= 0):
    #     load_value["req"]["operationTypeEnum"] = "NULL"
    #     sku["useQty"] = sku["holdQty"]
    #     sku["holdQty"] = 0
    #     use_qty_str[row_sku_str] += sku["useQty"]
    #     use_qty_str_log[row_sku_str] += sku["useQty"]
    # else:
    #     continue

    load_value = load_value["req"]
    # load_value = json.dumps(load_value)
    print(load_value)
    sheet.cell(row=i + 1, column=len(row) + 6, value=billNo)
    # 发送POST请求
    response = requests.post(url, json=load_value, headers=headers)
    print(response.text)
    # # # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    j+=1
    # if i>=1:
    #     break
f_tm= datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
print(j)
use_qty_str = json.dumps(use_qty_str)
print(use_qty_str)
use_qty_str_log = json.dumps(use_qty_str_log)
print(use_qty_str_log)
workbook.save(f_tm + '_tmp-wms.xlsx')