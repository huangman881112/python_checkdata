import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import commonUtil as cu


import requests
import json

# 目标URL
apiList = ''
with open('api_list', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        apiList += line
apiList = json.loads(apiList)
# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

warehouse_codes = """{"01": "上海集散仓", "2": "头程海运虚拟仓", "3": "上海待质检仓", "4": "上海不良品仓","5": "虚拟报废仓","06": "百世詹姆斯堡正品仓",
"10": "百世洛杉矶正品仓","11": "西雅图正品仓","7": "西雅图二手件仓","9": "西雅图配件仓","8": "西雅图报废仓","12": "FBA-01","13": "FBA-02",
"14": "FBA-03","15": "FBA-04","16": "FBA-05","17": "FBA-06","18": "FBA-07","19": "调拨虚拟仓","20": "FBA-04-CA",
"21": "华南工厂虚拟仓","22": "华东工厂虚拟仓","23": "新泽西正品仓","24": "FBA-08","25": "FBA-09","26": "YITALLC","27": "头程空运虚拟仓",
"28": "FBA-10","29": "无忧达亚特兰大","30": "上海不良品业务仓","31": "FBA-11","32": "FBA-12","33": "FBA-13","34": "FBA-14",
"35": "新泽西二手件仓","36": "上海集散仓-样品仓","37": "西雅图-不良品仓","38": "西雅图二手件-不良品仓","39": "新泽西-不良品仓","40": "FBA-15",
"41": "FBA-16","42": "FBA-17","43": "FBA-18","44": "FBA-19","45": "百世詹姆斯堡-不良品仓","46": "百世洛杉矶-不良品仓",
"47": "无忧达亚特兰大-不良品仓","48": "FBA-20","49": "FBA-21","50": "亚特兰大美南正品仓","51": "新泽西二手件-不良品仓","52": "亚特兰大美南二手件",
"53": "亚特兰大美南-不良品仓","54": "洛杉矶正品仓","55": "洛杉矶不良品仓","56": "FBA-22-UK","57": "WFS-01","58": "佛山集散仓","59": "佛山不良品仓",
"60": "佛山不良品业务仓","61": "NJ正品仓","62": "NJ-不良品仓","63": "TX正品仓","64": "TX-不良品仓","65": "FBA-22-EU","66": "FBA-01-DI",
"67": "4PX-德国法兰克福正品仓","68": "4PX-德国法兰克福-不良品仓","69": "4PX-英国莱切斯特正品仓","70": "4PX-英国莱切斯特-不良品仓","71": "CG-01",
"72": "WFS-02","73": "FBA-26-UK","74": "FBA-26-EU","75": "FBA-01-EU-DI","76": "FBA-01-UK-DI","77": "FBA-25-EU",
"78": "FBA-25-UK","79": "CG-02","80": "加州2号仓","81": "加州2号仓不良品仓","82": "FBA-01-EU","83": "FBA-01-UK","84": "ONUS-ONT",
"85": "ONUS-ONT-DS","86": "ONUS-CHINO","87": "ONUS-CHINO-DS","88": "ONUS-NJ-001","89": "ONUS-NJ-001-DS",
"90": "ONUS-NJ-002","91": "ONUS-NJ-002-DS","92": "ONUS-GA","93": "ONUS-GA-DS","94": "Dewell-NJ","95": "Dewell-NJ-DS",
"96": "德州2号仓","97": "德州2号仓不良品仓","98": "龙辕-CA","99": "龙辕-NJ","100": "龙辕-CA-DS","101": "龙辕-NJ-DS","102": "DUKE-ONT",
"103": "DUKE-ONT-DS","104": "Best-CA","105": "Best-CA-DS","106": "Best-NJ","107": "Best-NJ-DS","108": "中链英国仓",
"109": "中链英国仓-DS","110": "加州3号仓","111": "加州3号仓不良品仓","112": "亚特兰大2号仓","113": "亚特兰大2号仓不良品仓","114": "FBA-02-CA",
"115": "谷仓德国3仓良品仓","116": "谷仓德国3仓不良品仓","117": "谷仓英国16仓良品仓","118": "谷仓英国16仓不良品仓","119": "谷仓德国1号仓良品仓",
"120": "谷仓德国1号仓不良品仓","121": "FBA-27-UK","122": "FBA-27-EU","123": "谷仓德国7号仓良品仓","124": "谷仓德国7号仓不良品仓"}"""

warehouse_codes = json.loads(warehouse_codes)
filename = "d:\\tmp\\sku_dif\\执行结果0712-sku_dif_平台.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

# content = ''
# with open(tmpfile, 'r', encoding='utf-8') as file:
#     # content = file.read()  # 一次性读取整个文件内容
#     # 或者使用以下方式逐行读取
#     for line in file:
#         content += line
#     print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件

def convert_to_csv(filename,type,warehouse_code):
    # 选择工作表
    workbook = load_workbook(filename)
    sheet = workbook.active
    i = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        i = i + 1
        if row[0] is None:
            break

        if i == 1:
            sheet.cell(1, len(row) + 1, "sku_code")
            sheet.cell(1, len(row) + 2, "old_sku_code")
            sheet.cell(1, len(row) + 3, "warehouse_name")
            sheet.cell(1, len(row) + 4, "warehouse_code")
            sheet.cell(1, len(row) + 5, "platform")
            sheet.cell(1, len(row) + 6, "store")
            sheet.cell(1, len(row) + 7, "dif_sku")
            sheet.cell(1, len(row) + 8, "99_occpy")
            sheet.cell(1, len(row) + 9, "wms_storage_inv")
        else:
            new_value = row[2]
            str_list = new_value[19:].split(',')
            logic_qty = int(row[3])
            ordes_qty = int(row[4])
            sku_code = str_list[0]
            sku_dif = logic_qty - ordes_qty
            sheet.cell(i, len(row) + 1, str_list[0])
            sheet.cell(i, len(row) + 2, str_list[1])
            if warehouse_code is not None and warehouse_code != str_list[2]:
                continue
            if str_list[2] in warehouse_codes:
                sheet.cell(i, len(row) + 3, warehouse_codes[str_list[2]])
                sheet.cell(i, len(row) + 4, str_list[2])
                if type == "实物":
                    skustock = cu.getInvstockbySku(sku_code, str_list[2])
                    sheet.cell(i, len(row) + 5, skustock["platform"])
                    sheet.cell(i, len(row) + 6, skustock["store"])
                    sheet.cell(i, len(row) + 8, skustock["holdQty"])
                    wmsstorageinv = cu.getWmsStoregeStock(sku_code, str_list[2],"int")
                    sheet.cell(i, len(row) + 9, wmsstorageinv)
            sheet.cell(i, len(row) + 7, sku_dif)
        # row[2]=str_list[0]

        # load_value = json.loads(new_value)
        # load_value["req"]["remark"] = load_value["req"]["billNo"]
        # tm = load_value["req"]["billTime"]
        # print(tm)
        # tm = time.localtime(tm/1000)
        #
        # tm = time.strftime('%Y-%m-%d %H:%M:%S', tm)
        # load_value["req"]["billTime"]= tm
        # print(load_value)
        # load_value = json.dumps(load_value)
        # 发送POST请求
        # response = requests.post(url, json=load_value, headers=headers)
        # print(response.text)
        # 检查响应状态码
        # if response.status_code == 200:
        #     print("请求成功")
        #     # 处理响应数据，这里以JSON为例
        #     response_data = response.json()
        #     print(response_data)
        # else:
        #     print(f"请求失败，状态码：{response.status_code}")
    f_tm = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    print(i)
    if warehouse_code is None:
        warehouse_code = "all"
    workbook.save(f_tm +'_'+warehouse_code+ '_skudif.xlsx')


convert_to_csv(filename,"PINGTAI",None)