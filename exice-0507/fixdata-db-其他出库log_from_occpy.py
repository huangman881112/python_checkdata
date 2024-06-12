from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook
import commonUtil as comutil

import pandas as pd

import requests
import json

# 目标URL
# url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)
# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}


req_exampl = """
{
	"billNo": "RD24052914162",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-29T16:29:42.289",
	"billTypeEnum": "OTHER_OUTBOUND",
	"operationTypeEnum": "HOLD",
	"skuList": [
		{
			"badHoldQty": -1,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": 0,
			"inTransitQty": 0,
			"oldSkuCode": "BFTLPT-1011",
			"platform": "10001",
			"skuCode": "BFTLPT-1011",
			"store": "10415",
			"totalQty": -1,
			"useQty": 0
		}
	],
	"warehouseCode": "37"
}		
		

"""




filename = "d:\\tmp\\RD24052914474.xlsx"

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)
apilist = comutil.getApiList()
sku_List = comutil.getskuList()
# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    skuCode = row[9]
    oldSkuCode = row[10]
    platform = row[7]
    store = row[8]
    billNo = row[18]
    origin_bill_no = row[18]
    bill_time = row[15]
    warehouse_code = row[6]
    row_sku_str = skuCode + oldSkuCode + platform + store + warehouse_code
    print(row_sku_str)
    # print(sku_List)
    if row_sku_str not in sku_List:
        continue
    cur_qty = int(row[11])
    # qty = int(row[20])
    update_time = row[15]
    req_exam =comutil.init_data(req_exampl,None)
    req_exam["billNo"] = billNo
    req_exam["remark"] = billNo
    req_exam["originBillNo"] = billNo
    # print(type(update_time))
    print(update_time)
    updateTime = datetime.fromisoformat(update_time)
    # update_time = datetime.strftime(update_time,"%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    req_exam["warehouseCode"] = warehouse_code
    req_exam["operationTypeEnum"] = "NULL"
    skuList = req_exam["skuList"]
    for sku in skuList:
        sku["platform"] = platform
        sku["store"] = store
        sku["skuCode"] = skuCode
        sku["oldSkuCode"] = oldSkuCode
        sku["badHoldQty"] = cur_qty * -1
        sku["totalQty"] = cur_qty * -1

    print(req_exam)
    time.sleep(0.5)
    # 发送POST请求
    # response = requests.post(apilist["stock_log_url"], json=req_exam, headers=headers)
    # print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i>=1:
    #     break
f_tm= datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(i)
# workbook.save(f_tm + 'example.xlsx')