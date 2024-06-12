from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStockLog'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)

# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str  # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

billlist = """
2404H-G31-NJ
"""


head_trans_out = """
{
	"operationTypeEnum": "NULL",
	"skuList": [
		{
			"freezeQty": 0,
			"holdQty": -40,
			"useQty": 0,
			"totalQty": 0,
			"inTransitQty": 0,
			"store": "10415",
			"platform": "10001",
			"badQty": 0,
			"oldSkuCode": "FTHKCT-0005",
			"skuCode": "FTHKCT-0005",
			"badHoldQty": 0
		}
	],
	"originBillNo": "2404H-G200-CA03",
	"warehouseName": "佛山集散仓",
	"cooperator": "赢他",
	"billTypeEnum": "WMS_STRAIGHT_WAREHOUSRE",
	"warehouseCode": "58",
	"billTime": 1715926237269,
	"whType": "oneself",
	"billNo": "2404H-G200-CA03",
	"billStatusEnum": "FINISH"
}
"""


def init_data(rma_exampl,ignored_data):
    result = {}
    rma_exampl = json.loads(rma_exampl)
    for data_pro in rma_exampl:
        if data_pro is not None and data_pro == ignored_data:
            continue
        result[data_pro] = rma_exampl[data_pro]
    # print(result)
    return result


sku_str_list = ''
with open('sku_list_log_0', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        sku_str_list += line
    print(sku_str_list)

filename = "d:\\tmp\\tmp-stocklog.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)

# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)
billNo = ''
# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    load_value = row[14]
    print(load_value)
    load_value = json.loads(load_value)
    req_param = load_value["req"]
    billNo = req_param["billNo"]
    # if billNo not in billlist:
    #     continue
    warehouseCode = req_param["warehouseCode"]
    billTime = req_param["billTime"]
    sku_list = req_param["skuList"]
    req_param = json.dumps(req_param)
    req_exam = init_data(req_param,"skuList")
    skulist =[]
    for sku in sku_list:
        skuCode = sku["skuCode"]
        oldSkuCode = sku["oldSkuCode"]
        platform = sku["platform"]
        store = sku["store"]
        # skustr = skuCode + oldSkuCode + platform + store + warehouseCode
        # if skustr not in sku_str_list:
        #     continue
        sku= json.dumps(sku)
        sku_exam = init_data(sku,None)

        sku_exam["holdQty"] = sku_exam["holdQty"]*-1
        sku_exam["totalQty"] =  sku_exam["holdQty"]

        skulist.append(sku_exam)
    # print(type(update_time))

    req_exam["skuList"] = skulist
    req_exam["operationTypeEnum"] = "NULL"
    req_exam["remark"] = billNo

    # print(billTime)
    updateTime = datetime.fromtimestamp(billTime/1000)
    # update_time = datetime.strftime(updateTime, "%Y-%m-%d %H:%M:%S")
    req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
    # req_exam = json.dumps(req_exam)
    print(req_exam)
    #发送POST请求

    response = requests.post(url, json=req_exam, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
print(i)
workbook.save(f_tm + '_' + billNo + '_headtrans-出运.xlsx')
