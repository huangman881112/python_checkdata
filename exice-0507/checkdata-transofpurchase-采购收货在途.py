from datetime import datetime
import secrets
import string
import time

from openpyxl import Workbook
from openpyxl import load_workbook
import commonUtil as commutils

import requests
import json

biz_type_list = []


def getBillLists(url, body, startPage, pageSize, status, bill_list):
    print(url)
    suplr_strt_para_body = commutils.getApiList()[body]
    if commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY_STATUS in suplr_strt_para_body:
        suplr_strt_para_body[commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY_STATUS] = status
    elif commutils.HEAD_DEPOTRESERVATION_PARA_BODY_COMPLETESTATUS in suplr_strt_para_body:
        suplr_strt_para_body[commutils.HEAD_DEPOTRESERVATION_PARA_BODY_COMPLETESTATUS] = status
    suplr_strt_para_body[commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY_PAGESIZE] = pageSize
    suplr_strt_para_body[commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY_CURRENTPAGE] = startPage
    suplr_strt_response = requests.post(commutils.getApiList()[url],
                                        json=suplr_strt_para_body, headers=commutils.getHeaders())
    suplr_strt_res = json.loads(suplr_strt_response.text)
    # print(suplr_strt_res[commutils.DATA])
    for record in suplr_strt_res[commutils.DATA][commutils.RECORDS]:
        if "planNo" in record:
            bill_list.append(record["planNo"])
        elif "reservationID" in record:
            bill_list.append(record["reservationID"])
    startPage += 1
    while suplr_strt_res[commutils.DATA]["total"] > (startPage - 1) * pageSize:
        print(startPage)
        return getBillLists(url,body,startPage, pageSize, status, bill_list)

    return bill_list


bill_nolist = getBillLists(commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_URL, commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY,
                           1, 100, 1, [])
bill_nolist = getBillLists(commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_URL, commutils.HEAD_SUPPLYSTRAIGHT_LIST_ALL_PARA_BODY,
                           1, 100, 0, bill_nolist)
bill_nolist = getBillLists(commutils.HEAD_DEPOTRESERVATION_LIST_URL, commutils.HEAD_DEPOTRESERVATION_PARA_BODY,
                           1, 100, 0, bill_nolist)
bill_nolist = getBillLists(commutils.HEAD_DEPOTRESERVATION_LIST_URL, commutils.HEAD_DEPOTRESERVATION_PARA_BODY,
                           1, 100, 1, bill_nolist)


print(len(bill_nolist))

result = ''
with open('checkdata-transofpurchase-采购收货在途.txt', 'w', encoding='utf-8') as filewr:
    for bill_no in bill_nolist:
        inv_param_body = commutils.getApiList()[commutils.INV_OCCPY_PARA_BODY]

        requests.post(commutils.getApiList()[commutils.INV_OCCPY_URL], json=param_body,
                                     headers=commutils.getHeaders())
        qty = int(bill_nolist[bill_no])
        # print(qty)
        if bill_no.startswith("24"):
            param_body = commutils.getApiList()[commutils.HEAD_SUPPLYSTRAIGHT_LIST_PARA_BODY]
            param_body["planNo"] = bill_no
            response = requests.post(commutils.getApiList()[commutils.HEAD_SUPPLYSTRAIGHT_LIST_URL], json=param_body,
                                     headers=commutils.getHeaders())
        else:
            param_body = commutils.getApiList()[commutils.HEAD_DEPOTRESERVATION_PARA_BODY]
            param_body["reservationID"] = bill_no
            response = requests.post(commutils.getApiList()[commutils.HEAD_DEPOTRESERVATION_LIST_URL], json=param_body,
                                     headers=commutils.getHeaders())
        print(response.text)
        time.sleep(0.5)
        response_data = json.loads(response.text)
        response_records = response_data["data"]
        #直发查询
        if 'status' in response_records:
            if response_records["status"] in [0, 1] and 'countQty' in response_records and response_records[
                "countQty"] == qty:
                print(bill_no)
            else:
                bill_str = bill_no + "," + (str(qty) + "\n")
                filewr.write(bill_str)
            continue
        #预约进仓查询
        elif 'records' in response_records and len(response_records["records"]) > 0:
            if response_records["records"][0]["completeStatus"] in [0, 1] and response_records["records"][0][
                "totalDeliveryQty"] == qty:
                print(bill_no)
            else:
                bill_str = bill_no + "," + (str(qty) + "\n")
                filewr.write(bill_str)
            continue
        bill_str = bill_no + "," + (str(qty) + "\n")
        print(bill_str)
        filewr.write(bill_str)
        rec_str = json.dumps(response_records)
        # print(bill_no,qty)

# 目标URL
url = ''

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }

# 请求头信息


head_out = """
{
	"billNo": "2406H-Q61-NJ",
	"billStatusEnum": "FINISH",
	"billTime": "2024-05-31T13:18:43",
	"billTypeEnum": "TMS_STRAIGHT_WAREHOUSRE",
	"operationTypeEnum": "PUR_RECEIVE_IN_TRANSIT",
	"originBillNo": "ZPO24032500026",
	"skuList": [
		{
			"badHoldQty": 0,
			"badQty": 0,
			"freezeQty": 0,
			"holdQty": 0,
			"inTransitQty": 98,
			"oldSkuCode": "BFTLPT-1002",
			"platform": "10001",
			"skuCode": "BFTLPT-1002",
			"store": "10415",
			"totalQty": 0,
			"useQty": 0
		}
	],
	"warehouseCode": "01"
}

"""

#
# def init_data(rma_exampl):
#     result = {}
#     rma_exampl = json.loads(rma_exampl)
#     for data_pro in rma_exampl:
#         result[data_pro] = rma_exampl[data_pro]
#     # print(result)
#     return result
#
#
# def geneare_str(length):
#     characters = string.ascii_letters + string.digits
#     random_string = ''.join(secrets.choice(characters) for _ in range(length))
#     return random_string
#
#
# filename = "d:\\tmp\\占用差异-0513.xlsx"
# tmpfile = "d:\\tmp\\tmp.txt"
#
# content = ''
# with open(tmpfile, 'r', encoding='utf-8') as file:
#     # content = file.read()  # 一次性读取整个文件内容
#     # 或者使用以下方式逐行读取
#     for line in file:
#         content += line
#     # print(content)
#
# # df = pd.read_excel(filename)
# i = 0
# # 加载现有的Excel文件
# workbook = load_workbook(filename)
# billNo = ''
# # 选择工作表
# sheet = workbook.active
# for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
#     if row[0] is None:
#         break
#     biz_type = row[1]
#     if biz_type != 'GOOD_OUT_HOLD':
#         continue
#     skuCode = row[9]
#     oldSkuCode = row[10]
#     platform = row[7]
#     store = row[8]
#     # billNo = row[19]
#     # origin_bill_no = row[18]
#     bill_time = row[15]
#     warehouse_code = row[6]
#     qty = int(row[11])
#     update_time = row[15]
#     req_exam = init_data(head_out)
#     gentime = datetime.now().strftime('%m%d%H')
#     if billNo == '':
#         billNo = gentime + geneare_str(4) + "_invfix"
#     req_exam["billNo"] = billNo
#     req_exam["remark"] = billNo
#     req_exam["originBillNo"] = billNo
#     # print(type(update_time))
#     print(update_time)
#     updateTime = datetime.fromisoformat(update_time)
#     # update_time = datetime.strftime(update_time,"%Y-%m-%d %H:%M:%S")
#     req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
#     req_exam["warehouseCode"] = warehouse_code
#     skuList = req_exam["skuList"]
#     for sku in skuList:
#         sku["platform"] = platform
#         sku["store"] = store
#         sku["skuCode"] = skuCode
#         sku["oldSkuCode"] = oldSkuCode
#         sku["holdQty"] = qty * -1
#         sku["totalQty"] = 0
#     req_exam = json.dumps(req_exam)
#     print(req_exam)
#     # 发送POST请求
#
#     # print(response.text)
#     # 检查响应状态码
#     # if response.status_code == 200:
#     #     print("请求成功")
#     #     # 处理响应数据，这里以JSON为例
#     #     response_data = response.json()
#     #     print(response_data)
#     # else:
#     #     print(f"请求失败，状态码：{response.status_code}")
#     i += 1
#     if i > 1:
#         break
# f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
# print(i)
# workbook.save(f_tm +'_' +billNo+'_head_调出.xlsx')
