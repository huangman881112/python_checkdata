import Action
from datetime import datetime
import json
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import json
import commonUtil as comutils

import Action


def excute_errorlog(filename, head_trans_out, action: Action, m, flag):
    workbook = load_workbook(filename)
    sku_fixed = []
    # 选择工作表
    i = 0
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        new_value = row[14]
        skuCode = row[4]
        oldSkuCode = row[5]
        warehouseCode = row[6]
        rowskuCode = skuCode + oldSkuCode + warehouseCode
        load_value = json.loads(new_value)
        load_value["req"]["remark"] = load_value["req"]["billNo"]
        # print(load_value["req"]["billNo"])
        billno = load_value["req"]["billNo"]
        sheet.cell(i + 2, len(row) + 1, load_value["req"]["originBillNo"])
        tm = load_value["req"]["billTime"]
        tm = time.localtime(tm / 1000)
        tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
        load_value["req"]["billTime"] = tm
        # load_value["req"]["operationTypeEnum"]="NULL"
        load_value = load_value["req"]
        skulist = load_value["skuList"]
        newskuList = []
        for sku in skulist:
            _skuCode = sku["skuCode"]
            _oldSkuCode = sku["oldSkuCode"]
            if skuCode + oldSkuCode != _skuCode + _oldSkuCode:
                continue
            sku_str = json.dumps(sku)
            sku_exam = comutils.init_data(sku_str, None)
            newskuList.append(sku_exam)
        # 发送POST请求
        load_value["skuList"] = newskuList
        load_value = action.execute_log_req(load_value)
        if flag == False:
            load_value = json.dumps(load_value)
            print(load_value)
        if flag == True and load_value is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_LOG_URL], json=load_value,
                                     headers=comutils.getHeaders())
            print(response.text)
        sku_fixed.append(rowskuCode)
        i += 1
        if m > 0 and i >= m:
            break


action = {}


def excute_occpy(filename, head_trans_out, action: Action, m, flag):
    i = 0
    # 加载现有的Excel文件
    workbook = load_workbook(filename)
    billNo = ''
    # 选择工作表
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        sku_exam = {}
        skuCode = row[9]
        sku_exam["skuCode"] = skuCode
        oldSkuCode = row[10]
        sku_exam["oldSkuCode"] = oldSkuCode
        platform = row[7]
        sku_exam["platform"] = platform
        store = row[8]
        sku_exam["store"] = store
        billNo = row[18]
        origin_bill_no = row[17]
        bill_time = row[15]
        warehouse_code = row[6]
        qty = int(row[11])
        update_time = bill_time
        req_exam = comutils.init_data(head_trans_out, None)
        req_exam["billNo"] = billNo
        req_exam["remark"] = billNo
        req_exam["originBillNo"] = origin_bill_no
        # print(skuCode + oldSkuCode + platform + store + warehouse_code)
        updateTime = datetime.fromisoformat(update_time)
        req_exam["billTime"] = updateTime.strftime("%Y-%m-%dT%H:%M:%S")
        req_exam["warehouseCode"] = warehouse_code
        skuList = req_exam["skuList"]
        for sku in skuList:
            for prop in sku:
                if sku[prop] == 0 or sku[prop] == '':
                    continue
                elif prop.endswith('Qty'):
                    sku[prop] = qty
                else:
                    sku[prop] = sku_exam[prop]

        # 发送POST请求
        req_exam["skuList"] = action.excute_sku(skuList)
        req_exam = action.excute_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
        # print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_LOG_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break



def execute_stock_log(filename, head_trans_out, action: Action, m, flag):
    i = 0
    # 加载现有的Excel文件
    workbook = load_workbook(filename)
    billNo = ''
    # 选择工作表
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        sku_exam = {}
        skuCode = row[0]
        sku_exam["skuCode"] = skuCode
        oldSkuCode = row[1]
        sku_exam["oldSkuCode"] = oldSkuCode
        platform = row[2]
        sku_exam["platform"] = platform
        store = row[3]
        sku_exam["store"] = store
        now = datetime.now()
        billNo = now.strftime("%Y%m%d")+"_inv_fixlog"
        # billNo = row[18]
        # origin_bill_no = row[17]
        # bill_time = row[15]
        warehouse_code = row[4]
        qty = int(row[5])
        if qty > 0:
            head_trans_out = comutils.OTHER_INBOUND
        else:
            head_trans_out = comutils.OTHER_OUTBOUND
        req_exam = comutils.init_data(head_trans_out, None)
        req_exam["billNo"] = billNo
        req_exam["remark"] = billNo
        req_exam["originBillNo"] = billNo
        # print(skuCode + oldSkuCode + platform + store + warehouse_code)
        # updateTime = datetime.fromisoformat(update_time)
        req_exam["billTime"] = now.strftime("%Y-%m-%dT%H:%M:%S")
        req_exam["warehouseCode"] = warehouse_code
        skuList = req_exam["skuList"]
        for sku in skuList:
            for prop in sku:
                if sku[prop] == 0 or sku[prop] == '':
                    continue
                elif prop.endswith('Qty'):
                    sku[prop] = qty
                else:
                    sku[prop] = sku_exam[prop]

        # 发送POST请求
        req_exam = action.execute_log_req(req_exam)
        if flag == False:
            req_exam = json.dumps(req_exam)
            print(req_exam)
        if flag == True and req_exam is not None:
            response = requests.post(comutils.getApiList()[comutils.STOCK_LOG_URL], json=req_exam,
                                     headers=comutils.getHeaders())
            print(response.text)
        i += 1
        if m > 0 and i >= m:
            break
    return i





def excute_sku(action: Action):
    action.handle_sku()


def excute_req(action: Action):
    action.handle_req()
