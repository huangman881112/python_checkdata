from datetime import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}

filename = "d:\\tmp\\生产在途差异0527-5.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)


# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)
po_zip_list={}
# 选择工作表
sheet = workbook.active
f_tm = datetime.now().strftime('%Y-%m-%d_%H%M%S')
with open('purchases_生产在途_po.txt','w',encoding='utf-8') as filewrite:
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is None:
            break
        origin_bill_no = row[2]
        print(origin_bill_no)
        for ori_bill in origin_bill_no.split(','):
            if ori_bill.startswith("PO"):
                origin_bill_no=ori_bill
                break;
        bill_no = row[3]
        # origin_bill_no = origin_bill_no+bill_no
        qty = int(row[11])
        print("origin_bill_no:",origin_bill_no,"bill_no:",bill_no,"qty:",qty)
        if origin_bill_no in po_zip_list:
            po_zip_list[origin_bill_no] += qty
        else:
            po_zip_list[origin_bill_no] = qty

        # new_value = row[14]
        # load_value = json.loads(new_value)
        # load_value["req"]["remark"] = load_value["req"]["billNo"]
        # # print(load_value["req"]["billNo"])
        # sheet.cell(i+2, len(row)+1, load_value["req"]["originBillNo"])
        # tm = load_value["req"]["billTime"]
        # tm = time.localtime(tm/1000)
        # tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
        # load_value["req"]["billTime"]= tm
        # # load_value["req"]["operationTypeEnum"]="NULL"
        # load_value = load_value["req"]
        # skulist = load_value["skuList"]
        # for sku in skulist:
        #     sku["holdQty"] = 0
        #     sku["totalQty"] = sku["useQty"]
        # load_value = json.dumps(load_value)
        # print(load_value)
        # 发送POST请求
        # response = requests.post(url, json=load_value, headers=headers)
        # print(response.text)
        # 检查响应状态码
        # if response.status_code == 200:
        #     print("请求成功")
        #     # 处理响应数据，这里以JSON为例
        #     response_data = response.json()
        #     print(response_data)
        # else:
        #     print(f"请求失败，状态码：{response.status_code}")
        i+=1
        # if i >=1:
        #     break
    filewrite.write(json.dumps(po_zip_list))

print(i)
# workbook.save(f_tm + filename[7:])