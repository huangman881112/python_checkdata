import json
from openpyxl import Workbook
from datetime import datetime

# filename = "d:\\tmp\\tmp.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)

# 加载现有的Excel文件
workbook = Workbook()

# 选择工作表
sheet = workbook.active
i = 0
j = 1
content = json.loads(content)

if 'data' in content:
    content = content['data']

if 'billNo' in content:
    bill_no = content['billNo']
elif 'billId' in content:
    bill_no = content['billId']

for data_pro in content:
    if data_pro == 'skuList':
        skulist = content[data_pro]
        m = 1
        for sku in skulist:
            i = 1
            for sku_pro in sku:
                if m == 1:
                    sheet.cell(1, i + j, sku_pro)
                    sheet.cell(2, i + j, sku[sku_pro])
                else:
                    sheet.cell(m+1, i + j, sku[sku_pro])
                i = i + 1
            m = m + 1
    else:
        sheet.cell(1, i + j, data_pro)
        sheet.cell(2, i + j, content[data_pro])
        j = j + 1

print(i)
# 保存对文件的更改

file_time = datetime.now().strftime('%Y%m%d_%H%M%S')
workbook.save(file_time+"_"+bill_no+"_"+'json2excel.xlsx')
# 关闭文件
workbook.close()
