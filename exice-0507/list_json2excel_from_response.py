import json
from openpyxl import Workbook
from datetime import datetime

import commonUtil as comutils



# filename = "d:\\tmp\\tmp.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    # print(content)

# 加载现有的Excel文件
workbook = Workbook()

# 选择工作表
sheet = workbook.active
i = 1
j = 1
l=0
bill_no = ''
content = json.loads(content)
if type(content["data"]) != list:
    if 'records' in content["data"]:
        content_data = content["data"]["records"]
    elif 'list' in content["data"]:
        content_data = content["data"]["list"]
else:
    content_data = content["data"]
for row_data in content_data:
    i = i + 1+l
    j = 1
    for data_pro in row_data:

        if row_data[data_pro] is None or row_data[data_pro] == '':
            continue
        if data_pro == 'batchNo' or data_pro == 'shipmentPlanId' or data_pro == 'billId' or data_pro == 'busCode':
            bill_no = row_data[data_pro]
        if i == 2:
            sheet.cell(1, j, data_pro)
            if type(row_data[data_pro]) == list:
                l = len(row_data[data_pro])
                r=0
                for item in row_data[data_pro]:
                    c=0
                    r+=1
                    if r==1:
                        for item_prop in item:
                            # print(item_prop)
                            sheet.cell(1, j +len(row_data) + c, item_prop)
                            sheet.cell(2, j +len(row_data)+ c, item[item_prop])
                            c+=1
                    else:
                        for item_prop in item:
                            sheet.cell(r+1, j+len(row_data)+c, item[item_prop])
                            c+=1
            else:
                sheet.cell(2, j, row_data[data_pro])
        else:
            if type(row_data[data_pro]) == list:
                l = len(row_data[data_pro])
                r = 0
                for item in row_data[data_pro]:
                    r += 1
                    c=0
                    for item_prop in item:
                        sheet.cell(r + i, j + len(row_data) +c, item[item_prop])
                        c += 1
            else:
                sheet.cell(i, j, row_data[data_pro])
        j = j + 1

print(i)
# 保存对文件的更改

file_time = comutils.FILE_PROFIX+ datetime.now().strftime('%Y%m%d_%H%M%S')
workbook.save(file_time + "_" + bill_no + "_" + 'list_json2excel.xlsx')
# 关闭文件
workbook.close()


def write_excel(list,r,c,sheet):
    r1=1
    c1 =1
    for sku in list:
        r1 +=1
        if r1==1:
            for prop in sku:
              sheet.cell(r,c+c1,sku)
              c1+=1

        else:
            for prop in sku:
                sheet.cell(r+r1, c + c1, sku)
                c1 += 1
    return  sheet
