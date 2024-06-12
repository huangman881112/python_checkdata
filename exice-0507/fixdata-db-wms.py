import datetime
import json
import time

from openpyxl import Workbook

from openpyxl import load_workbook

import pandas as pd

import requests
import json

# 目标URL
url = 'https://inv-web.yintaerp.com/api/inv/invWarehouseRatio/executeLogicStock'

# 请求体数据
# data = {
#     "key1": "value1",
#     "key2": "value2"
# }
token_str = ''
with open('token', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        token_str += line
    print(token_str)
# 请求头信息
headers = {
    "Content-Type": "application/json",  # 假设API期望接收JSON格式的数据
    "Authorization": token_str # 例如，使用Bearer Token进行身份验证
    # "Custom-Header": "CustomValue"  # 自定义请求头字段
}
use_qty_str = ''
with open('use_qty', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str += line
    # print(use_qty_str)

use_qty_str = json.loads(use_qty_str)

use_qty_str_log = ''
with open('use_qty_log', 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        use_qty_str_log += line
    # print(use_qty_str)

use_qty_str_log = json.loads(use_qty_str_log)





filename = "d:\\tmp\\tmp-wms.xlsx"
tmpfile = "d:\\tmp\\tmp.txt"

content = ''
with open(tmpfile, 'r', encoding='utf-8') as file:
    # content = file.read()  # 一次性读取整个文件内容
    # 或者使用以下方式逐行读取
    for line in file:
        content += line
    print(content)


# df = pd.read_excel(filename)
i = 0
# 加载现有的Excel文件
workbook = load_workbook(filename)

# 选择工作表
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    if row[0] is None:
        break
    new_value = row[14]
    load_value = json.loads(new_value)
    load_value["req"]["remark"] = load_value["req"]["billNo"]
    tm = load_value["req"]["billTime"]
    # print(tm)
    tm = time.localtime(tm/1000)

    tm = time.strftime('%Y-%m-%dT%H:%M:%S', tm)
    load_value["req"]["billTime"]= tm
    sku_str = ''
    warehouseCode = load_value["req"]["warehouseCode"]
    sku_list =load_value["req"]["skuList"]
    for sku in sku_list:
        skucode= sku["skuCode"]
        oldskucode =sku["oldSkuCode"]
        platform = sku["platform"]
        store = sku["store"]
        sku_str += skucode + oldskucode + platform + store + warehouseCode
        print(sku_str)
        cur_qty = sku["holdQty"]
        sku["useQty"] = sku["holdQty"]
        sku["holdQty"] = 0
    if use_qty_str[sku_str]+cur_qty>=0:
        use_qty_str[sku_str] += cur_qty
        print(sku_str)
    else:
        continue
    load_value["req"]["operationTypeEnum"] = "NULL"
    load_value_req = load_value["req"]
    # load_value_req = json.dumps(load_value_req)
    print(load_value_req)
    # 发送POST请求
    response = requests.post(url, json=load_value_req, headers=headers)
    print(response.text)
    # 检查响应状态码
    # if response.status_code == 200:
    #     print("请求成功")
    #     # 处理响应数据，这里以JSON为例
    #     response_data = response.json()
    #     print(response_data)
    # else:
    #     print(f"请求失败，状态码：{response.status_code}")
    i+=1
    # if i >1:
    #     break
f_tm= datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
print(i)
workbook.save(f_tm + 'example.xlsx')